import asyncio
import threading
import struct
import time
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from collections import deque

from bleak import BleakClient, BleakScanner

# ── BLE UUID（与 myBle.py 一致） ───────────────────────────────────────────
BLE_TX_UUID        = "0000FFE4-0000-1000-8000-00805F9A34FB"  # 设备→PC Notify
BLE_RX_UUID        = "0000FFE9-0000-1000-8000-00805F9A34FB"  # PC→设备 Write
DEVICE_NAME_PREFIX = "Force"

# ── 帧格式（与 main.py / win.py 完全一致） ────────────────────────────────
FRAME_LEN         = 244
DATA_FIELD_LEN    = 240
MODBUS_SLAVE      = 0x01
MODBUS_FUNC       = 0x03
MODBUS_FUNC_MIXED = 0x05

# ── 指令（与 win.py 一致） ────────────────────────────────────────────────
CMD_RESTORE    = bytes([0x54, 0xF8])
CMD_SAVE_DEF   = bytes([0x54, 0xCC])
CMD_TEMP_REQ   = bytes([0x54, 0xDD])
CMD_DISCONNECT = bytes([0x54, 0xDE])
CMD_CALIB_HDR  = bytes([0x54, 0xAA])


class BLESensorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("测力数据采集 (BLE)")
        self.root.geometry("1280x720")
        self.root.minsize(900, 600)

        # BLE
        self.ble_client: BleakClient | None = None
        self.ble_loop:   asyncio.AbstractEventLoop | None = None
        self.ble_thread: threading.Thread | None = None
        self.running  = False
        self.scanning = False

        # 数据
        self.data_lock   = threading.Lock()
        self.data_buffer = bytearray()
        self.latest_filtered_weight = None
        self.latest_left_weight = None
        self.latest_right_weight = None
        self.is_dual_mode = False
        self.zero_offset  = 0.0

        # 录制
        self.recording          = False
        self._excel_rows        = []
        self._excel_header      = []
        self._excel_dual        = False
        self.filename           = ""
        self.record_timestamp_ms = 1
        self.current_frequency_hz = None

        # 温度
        self.resp_buffer         = bytearray()
        self.pending_temp_request = False

        # 三点滑动窗口
        self.THRESHOLD_DETECT = 30.0
        self.THRESHOLD_KEEP   = 5.0
        self._raw_buf = []
        self._res_buf = []
        self._left_raw_buf = []
        self._left_res_buf = []
        self._right_raw_buf = []
        self._right_res_buf = []

        # 频率
        self.data_count       = 0
        self.last_print_time  = time.time()
        self.freq_print_running = False

        # 图表
        self.chart_max_points = 10000
        self.chart_data_x = deque(maxlen=self.chart_max_points)
        self.chart_data_y = deque(maxlen=self.chart_max_points)
        self.chart_data_left_y = deque(maxlen=self.chart_max_points)
        self.chart_data_right_y = deque(maxlen=self.chart_max_points)
        self.data_sequence = 0

        self._weight_update_id = None
        self._chart_update_id  = None

        self._build_ui()
        self._schedule_weight_update()
        self._schedule_chart_update()

        self.root.bind("r", self.handle_record_shortcut)
        self.root.bind("R", self.handle_record_shortcut)
        self.root.bind("s", self.handle_start_shortcut)
        self.root.bind("S", self.handle_start_shortcut)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ═══════════════════════════════════════════════════════════════════ UI
    def _build_ui(self):
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # ── 左：图表 ──────────────────────────────────────────────────────
        chart_frame = tk.Frame(main_frame)
        chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        top_bar = tk.Frame(chart_frame)
        top_bar.pack(side=tk.TOP, fill=tk.X, pady=(0, 6))
        weight_font = ("SimHei", 20, "bold")
        metric_pad = dict(side=tk.LEFT, padx=(0, 12))

        self.left_weight_label = tk.Label(top_bar, text="主设备: -- kg",
                                          font=weight_font, fg="#d32f2f",
                                          width=14, anchor="w")
        self.left_weight_label.pack(**metric_pad)
        self.right_weight_label = tk.Label(top_bar, text="副设备: -- kg",
                                           font=weight_font, fg="#1565c0",
                                           width=14, anchor="w")
        self.right_weight_label.pack(**metric_pad)
        self.weight_label = tk.Label(top_bar, text="合力: -- kg",
                                     font=weight_font, fg="#000000",
                                     width=14, anchor="w")
        self.weight_label.pack(**metric_pad)

        self.fig = Figure(dpi=100)
        self.ax  = self.fig.add_subplot(111)
        self.ax.set_xlabel("数据序号", fontproperties="SimHei")
        self.ax.set_ylabel("重量 (kg)", fontproperties="SimHei")
        self.ax.set_title("实时重量曲线", fontproperties="SimHei")
        self.ax.grid(True, alpha=0.3)
        self.line_left, = self.ax.plot([], [], color="#d32f2f", linewidth=1.4, label="主设备")
        self.line_right, = self.ax.plot([], [], color="#1565c0", linewidth=1.4, label="副设备")
        self.line_total, = self.ax.plot([], [], color="#000000", linewidth=1.6, label="合力")
        legend = self.ax.legend(prop={"family": "SimHei", "size": 10})
        for text, color in zip(legend.get_texts(), ["#d32f2f", "#1565c0", "#000000"]):
            text.set_color(color)
        self.fig.tight_layout()

        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        footer_bar = tk.Frame(chart_frame)
        footer_bar.pack(side=tk.BOTTOM, fill=tk.X, pady=(4, 0))
        self.freq_label = tk.Label(footer_bar, text="采集频率：-- Hz",
                                   fg="#777777", font=("SimHei", 10), anchor="w")
        self.freq_label.pack(side=tk.LEFT)
        self.mode_label = tk.Label(footer_bar, text="模式：--",
                                   fg="#777777", font=("SimHei", 10), anchor="e")
        self.mode_label.pack(side=tk.RIGHT)

        # ── 右：控制面板 ──────────────────────────────────────────────────
        right = tk.Frame(main_frame, width=220, bg="#f4f4f4",
                         relief=tk.GROOVE, bd=1)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(6, 0))
        right.pack_propagate(False)
        pad = dict(padx=10, pady=4, fill=tk.X)

        # 连接
        self._section(right, "连接")
        self.status_label = tk.Label(right, text="状态：未连接",
                                     bg="#f4f4f4", fg="#555", font=("SimHei", 9))
        self.status_label.pack(**pad)
        self.device_label = tk.Label(right, text="设备：--",
                                     bg="#f4f4f4", fg="#888", font=("SimHei", 8),
                                     wraplength=200, justify=tk.LEFT)
        self.device_label.pack(**pad)
        self.receive_button = tk.Button(right, text="扫描连接 (s)",
                                        command=self.toggle_receiving,
                                        bg="#4caf50", fg="white", font=("SimHei", 10),
                                        relief=tk.FLAT, bd=0, padx=10, pady=8,
                                        activeforeground="white", activebackground="#43a047",
                                        cursor="hand2")
        self.receive_button.pack(**pad)
        self.record_button = tk.Button(right, text="开始记录 (r)",
                                       command=self.toggle_recording,
                                       state=tk.DISABLED,
                                       bg="#2196f3", fg="white", font=("SimHei", 10),
                                       relief=tk.FLAT, bd=0, padx=10, pady=8,
                                       activeforeground="white", activebackground="#1e88e5",
                                       cursor="hand2")
        self.record_button.pack(**pad)
        plain_btn = dict(bg="#ebebeb", fg="#333333", activebackground="#dddddd",
                         activeforeground="#111111", font=("SimHei", 10),
                         relief=tk.FLAT, bd=0, padx=10, pady=8, cursor="hand2")

        # 归零
        self._section(right, "归零")
        tk.Button(right, text="归零（清零当前值）",
                  command=self.do_zero, **plain_btn).pack(**pad)
        tk.Button(right, text="取消归零",
                  command=self.cancel_zero, **plain_btn).pack(**pad)
        self.zero_label = tk.Label(right, text="偏移: 0.0 kg",
                                   bg="#f4f4f4", fg="#555", font=("SimHei", 9))
        self.zero_label.pack(**pad)

        # 校准
        self._section(right, "校准")
        tk.Label(right, text="砝码实际重量 (kg)",
                 bg="#f4f4f4", font=("SimHei", 9)).pack(**pad)
        calib_row = tk.Frame(right, bg="#f4f4f4")
        calib_row.pack(padx=10, pady=2, fill=tk.X)
        self.calib_actual_entry = tk.Entry(calib_row, width=10, font=("SimHei", 11))
        self.calib_actual_entry.insert(0, "1.000")
        self.calib_actual_entry.pack(side=tk.LEFT, padx=2)
        tk.Label(calib_row, text="kg", bg="#f4f4f4", font=("SimHei", 9)).pack(side=tk.LEFT)
        tk.Button(right, text="发送校准",
                  command=self.send_calibration, **plain_btn).pack(**pad)

        # 比例系数
        self._section(right, "比例系数")
        tk.Button(right, text="保存当前为默认值",
                  command=self.send_save_default, **plain_btn).pack(**pad)
        tk.Button(right, text="恢复默认值",
                  command=self.send_restore_default, **plain_btn).pack(**pad)

        # 温度
        self._section(right, "温度")
        tk.Button(right, text="读取温度原始值",
                  command=self.request_temperature, **plain_btn).pack(**pad)
        self.temp_label = tk.Label(right, text="温度原始值: --",
                                   bg="#f4f4f4", fg="#333", font=("SimHei", 10))
        self.temp_label.pack(**pad)

        tk.Frame(right, bg="#f4f4f4").pack(fill=tk.BOTH, expand=True)

    def _section(self, parent, title):
        frm = tk.Frame(parent, bg="#f4f4f4")
        frm.pack(fill=tk.X, padx=6, pady=(10, 2))
        tk.Label(frm, text=title, bg="#d0d0d0", fg="#333",
                 font=("SimHei", 9, "bold"), anchor="w",
                 relief=tk.FLAT, padx=4).pack(fill=tk.X)

    # ═══════════════════════════════════════════════════════════ 定时刷新
    def _schedule_weight_update(self):
        if self._weight_update_id:
            self.root.after_cancel(self._weight_update_id)
        self._weight_update_id = self.root.after(80, self._update_weight_display)

    def _update_weight_display(self):
        self._weight_update_id = None
        with self.data_lock:
            w     = self.latest_filtered_weight
            left  = self.latest_left_weight
            right = self.latest_right_weight
            dual  = self.is_dual_mode
        self.left_weight_label.config(text=f"左脚: {left} kg" if left is not None else "主设备: -- kg")
        self.right_weight_label.config(text=f"右脚: {right} kg" if right is not None else "副设备: -- kg")
        self.weight_label.config(text=f"合力: {w} kg" if w is not None else "合力: -- kg")
        self.mode_label.config(text="模式：双机（主设备+副设备=合力）" if dual else "模式：单机")
        self._schedule_weight_update()

    def _schedule_chart_update(self):
        if self._chart_update_id:
            self.root.after_cancel(self._chart_update_id)
        self._chart_update_id = self.root.after(100, self._update_chart)

    def _update_chart(self):
        self._chart_update_id = None
        with self.data_lock:
            x_data = list(self.chart_data_x)
            total_y = list(self.chart_data_y)
            left_y = list(self.chart_data_left_y)
            right_y = list(self.chart_data_right_y)

        n_total = min(len(x_data), len(total_y))
        n_left = min(len(x_data), len(left_y))
        n_right = min(len(x_data), len(right_y))

        self.line_total.set_data(x_data[-n_total:] if n_total > 0 else [],
                                 total_y[-n_total:] if n_total > 0 else [])
        self.line_left.set_data(x_data[-n_left:] if n_left > 0 else [],
                                left_y[-n_left:] if n_left > 0 else [])
        self.line_right.set_data(x_data[-n_right:] if n_right > 0 else [],
                                 right_y[-n_right:] if n_right > 0 else [])

        if n_total > 0 or n_left > 0 or n_right > 0:
            self.ax.relim()
            self.ax.autoscale_view()
            y_min, y_max = self.ax.get_ylim()
            if y_max <= 20:
                self.ax.set_ylim(y_min, 20)
            self.canvas.draw_idle()
        self._schedule_chart_update()

    # ═══════════════════════════════════════════════════════════════ 快捷键
    def handle_record_shortcut(self, event):
        if self.record_button["state"] != tk.DISABLED:
            self.toggle_recording()

    def handle_start_shortcut(self, event):
        self.toggle_receiving()

    # ═══════════════════════════════════════════════════════════════ BLE
    def toggle_receiving(self):
        if self.running or self.scanning:
            self._stop_ble()
        else:
            self._start_ble_thread()

    def _start_ble_thread(self):
        self.scanning = True
        self.receive_button.config(text="扫描中...", bg="#ff9800", state=tk.DISABLED)
        self.status_label.config(text="状态：正在扫描", fg="#e65100")

        def _run():
            self.ble_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(self.ble_loop)
            try:
                self.ble_loop.run_until_complete(self._ble_main())
            finally:
                self.ble_loop.close()
                self.ble_loop = None

        self.ble_thread = threading.Thread(target=_run, daemon=True)
        self.ble_thread.start()

    def _stop_ble(self):
        """从 tkinter 线程请求断开 BLE。"""
        self.running = False
        self.scanning = False
        if self.ble_loop and self.ble_loop.is_running():
            asyncio.run_coroutine_threadsafe(self._ble_disconnect(), self.ble_loop)
        else:
            self.root.after(0, self._on_ble_disconnected)

    async def _ble_disconnect(self):
        client = self.ble_client
        if client and client.is_connected:
            try:
                try:
                    await client.write_gatt_char(BLE_RX_UUID, CMD_DISCONNECT, response=True)
                    print("已发送主动断开指令")
                    await asyncio.sleep(0.2)
                except Exception as e:
                    print(f"发送主动断开指令失败: {e}")
                try:
                    await client.stop_notify(BLE_TX_UUID)
                except Exception:
                    pass
                await client.disconnect()
                print("已主动断开 BLE 连接")
            except Exception as e:
                print(f"BLE断开失败: {e}")

    async def _ble_main(self):
        """扫描 -> 连接 -> 接收 -> 断开。"""
        device = await self._scan_for_device()
        if device is None:
            self.root.after(0, lambda: self._on_ble_disconnected("未找到 Force 设备"))
            return

        dev_name = device.name or device.address
        self.root.after(0, lambda n=dev_name, a=device.address: (
            self.status_label.config(text=f"状态：连接中 {n}", fg="#e65100"),
            self.device_label.config(text=f"设备：{n}\n{a}"),
        ))

        try:
            await self._ble_connect(device)
            self.root.after(0, lambda: self._on_ble_disconnected())
        except Exception as e:
            self.freq_print_running = False
            self.running = False
            self.scanning = False
            self.ble_client = None
            self.root.after(0, lambda err=e: self._on_ble_disconnected(f"连接失败: {err}"))
        return

    async def _ble_connect(self, device):
        """单次连接尝试，由 _ble_main 调用。"""
        try:
            async with BleakClient(
                device,
                disconnected_callback=self._on_ble_disconnect_cb
            ) as client:
                self.ble_client = client
                self.running  = True
                self.scanning = False

                # 重置状态
                self.data_buffer = bytearray()
                self.resp_buffer = bytearray()
                self._raw_buf = []
                self._res_buf = []
                self.latest_filtered_weight = None
                self.is_dual_mode = False
                self.data_count = 0
                self.last_print_time = time.time()
                self.current_frequency_hz = None
                self.data_sequence = 0
                with self.data_lock:
                    self.chart_data_x.clear()
                    self.chart_data_y.clear()
                    self.chart_data_left_y.clear()
                    self.chart_data_right_y.clear()

                self.root.after(0, self._on_ble_connected)

                self.freq_print_running = True
                threading.Thread(target=self._print_frequency, daemon=True).start()

                await client.start_notify(BLE_TX_UUID, self._on_notify)

                while self.running and client.is_connected:
                    await asyncio.sleep(0.2)

                try:
                    await client.stop_notify(BLE_TX_UUID)
                except Exception:
                    pass
        except Exception as e:
            raise  # 抛给 _ble_main 重试逻辑
        finally:
            self.freq_print_running = False
            self.running  = False
            self.scanning = False
            self.ble_client = None

    async def _scan_for_device(self):
        print(f"扫描 BLE 设备（前缀: {DEVICE_NAME_PREFIX}）...")
        devices = await BleakScanner.discover(timeout=10.0)
        for d in devices:
            if d.name and d.name.startswith(DEVICE_NAME_PREFIX):
                print(f"找到设备: {d.name}  {d.address}")
                return d
        print("未发现 Force 设备，已发现的设备列表：")
        for d in devices:
            print(f"  {d.name or '(无名称)'}  {d.address}")
        return None

    # ── BLE 回调（在 asyncio 线程中调用） ────────────────────────────────
    def _on_notify(self, sender, data: bytearray):
        """BLE Notify 回调：优先解析控制响应，剩余字节再按数据帧解析。"""
        self.resp_buffer.extend(data)
        self._parse_responses()
        if self.resp_buffer:
            self.data_buffer.extend(self.resp_buffer)
            self.resp_buffer = bytearray()
        self.parse_data()

    def _on_ble_disconnect_cb(self, client: BleakClient):
        """bleak 断开回调（asyncio 线程），通知主线程。"""
        self.running = False
        self.root.after(0, self._on_ble_disconnected)

    # ── tkinter 线程回调 ─────────────────────────────────────────────────
    def _on_ble_connected(self):
        self.receive_button.config(text="断开 (s)", bg="#e53935", state=tk.NORMAL)
        self.record_button.config(state=tk.NORMAL)
        self.status_label.config(text="状态：接收数据中", fg="#2e7d32")
        self.freq_label.config(text="采集频率：-- Hz")

    def _on_ble_disconnected(self, reason: str = ""):
        self.receive_button.config(text="扫描连接 (s)", bg="#4caf50", state=tk.NORMAL)
        self.record_button.config(state=tk.DISABLED)
        self.freq_label.config(text="采集频率：-- Hz")
        self.current_frequency_hz = None
        msg = f"状态：未连接{' (' + reason + ')' if reason else ''}"
        self.status_label.config(text=msg, fg="#555")
        if self.recording:
            self.stop_recording()
        with self.data_lock:
            self.latest_filtered_weight = None

    # ═══════════════════════════════════════════════════════════ 指令发送
    def _send_cmd(self, data: bytes) -> bool:
        """向设备发送指令，BLE 连接未建立时提示。"""
        if not self.ble_client or not self.ble_client.is_connected:
            messagebox.showwarning("未连接", "请先扫描连接设备后再操作。")
            return False
        if not self.ble_loop:
            return False
        async def _write():
            try:
                await self.ble_client.write_gatt_char(BLE_RX_UUID, data, response=False)
            except Exception as e:
                print(f"BLE发送失败: {e}")
        asyncio.run_coroutine_threadsafe(_write(), self.ble_loop)
        return True

    def send_calibration(self):
        if not messagebox.askyesno("确认校准", "确认按当前砝码重量发送校准参数吗？"):
            return
        try:
            actual = float(self.calib_actual_entry.get())
            if actual == 0:
                messagebox.showerror("输入错误", "砝码重量不能为0")
                return
            with self.data_lock:
                current = self.latest_filtered_weight
            if current is None:
                messagebox.showerror("无数据", "当前无实时重量数据，请先连接设备并放上砝码")
                return
            factor = current / actual
            factor_int = int(round(factor * 1000))
            if factor_int < 1 or factor_int > 65535:
                messagebox.showerror("输入错误", f"校准因子超出范围: {factor_int}")
                return
            cmd = CMD_CALIB_HDR + struct.pack('>H', factor_int)
            if self._send_cmd(cmd):
                print(f"已发送校准因子: {factor:.4f} (整数={factor_int})")
                messagebox.showinfo("校准", f"校准因子已发送: {factor:.4f}")
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的数字")

    def send_restore_default(self):
        if not messagebox.askyesno("确认恢复默认", "确认恢复默认比例系数吗？"):
            return
        if self._send_cmd(CMD_RESTORE):
            print("已发送恢复默认值指令")
            messagebox.showinfo("恢复默认", "已发送恢复默认值指令")

    def send_save_default(self):
        if not messagebox.askyesno("确认保存默认", "确认将当前比例系数保存为默认值吗？"):
            return
        if self._send_cmd(CMD_SAVE_DEF):
            print("已发送保存默认值指令")
            messagebox.showinfo("保存默认", "已发送保存当前为默认值指令")

    def request_temperature(self):
        if self._send_cmd(CMD_TEMP_REQ):
            print("已发送温度请求")
            self.pending_temp_request = True
            self.temp_label.config(text="温度原始值: 等待中...")

    def do_zero(self):
        with self.data_lock:
            current = self.latest_filtered_weight
        if current is None:
            messagebox.showwarning("归零", "当前无数据，无法归零")
            return
        self.zero_offset += current
        self.zero_label.config(text=f"偏移: {self.zero_offset:.1f} kg")
        print(f"归零：偏移量设为 {self.zero_offset:.1f} kg")

    def cancel_zero(self):
        self.zero_offset = 0.0
        self.zero_label.config(text="偏移: 0.0 kg")
        print("已取消归零")

    # ═══════════════════════════════════════════════════════ 频率统计
    def _print_frequency(self):
        while self.freq_print_running and self.running:
            time.sleep(2)
            if self.freq_print_running and self.running:
                now = time.time()
                elapsed = now - self.last_print_time
                if elapsed > 0:
                    freq = self.data_count / elapsed
                    self.current_frequency_hz = freq
                    self.root.after(0, lambda f=freq: self.freq_label.config(text=f"采集频率：{f:.2f} Hz"))
                    print(f"采集频率: {freq:.2f} Hz (共 {self.data_count} 个数据点)")
                    self.data_count = 0
                    self.last_print_time = now

    # ═══════════════════════════════════════════════════════ 数据解析
    def _parse_responses(self):
        """从 resp_buffer 解析温度响应（0x54 0xDD + 4字节有符号整数）。"""
        buf = self.resp_buffer
        while len(buf) >= 6:
            found = False
            for i in range(len(buf) - 1):
                if buf[i] == 0x54 and buf[i + 1] == 0xDD:
                    if i + 6 <= len(buf):
                        raw = struct.unpack('>i', bytes(buf[i + 2:i + 6]))[0]
                        print(f"温度原始值: {raw}")
                        self.root.after(0, lambda v=raw: self.temp_label.config(
                            text=f"温度原始值: {v}"))
                        self.pending_temp_request = False
                        buf = buf[i + 6:]
                        found = True
                        break
                    else:
                        break
            if not found:
                break
        self.resp_buffer = buf

    def _three_bytes_to_weight_kg(self, high_byte, low_byte, checksum):
        expected = (high_byte + low_byte) & 0xFF
        if checksum != expected:
            print(f"警告：校验和不匹配 (期望:{expected:02X}, 实际:{checksum:02X})")
            return None
        weight_int = (high_byte << 8) | low_byte
        if weight_int >= 32768:
            weight_int -= 65536
        return weight_int / 10.0

    def _parse_uart_weight(self, high_byte, low_byte, checksum):
        if high_byte == 0 and low_byte == 0 and checksum == 0:
            return None
        expected = (high_byte + low_byte) & 0xFF
        if checksum != expected:
            print(f"警告：UART校验和不匹配 (期望:{expected:02X}, 实际:{checksum:02X})")
            return None
        weight_int = (high_byte << 8) | low_byte
        if weight_int >= 32768:
            weight_int -= 65536
        return weight_int / 10.0

    def parse_data(self):
        """解析 244 字节帧，逻辑与 win.py 完全一致。"""
        while len(self.data_buffer) >= FRAME_LEN:
            frame = bytes(self.data_buffer[:FRAME_LEN])
            if frame[0] != MODBUS_SLAVE or (
                    frame[1] != MODBUS_FUNC and frame[1] != MODBUS_FUNC_MIXED):
                print(f"帧头错误，跳过1字节。[{frame[0]:02X} {frame[1]:02X}]")
                self.data_buffer = self.data_buffer[1:]
                continue
            self.data_buffer = self.data_buffer[FRAME_LEN:]

            is_mixed = (frame[1] == MODBUS_FUNC_MIXED)
            with self.data_lock:
                self.is_dual_mode = is_mixed

            if is_mixed:
                for i in range(0, DATA_FIELD_LEN, 6):
                    left_kg = self._three_bytes_to_weight_kg(
                        frame[4+i], frame[5+i], frame[6+i])
                    right_kg = self._three_bytes_to_weight_kg(
                        frame[7+i], frame[8+i], frame[9+i])
                    if left_kg is None or right_kg is None:
                        continue
                    combined = left_kg + right_kg
                    self._push_point(combined, left_kg, right_kg)
                    if self.recording:
                        filtered_combined = self.latest_filtered_weight
                        filtered_left = self.latest_left_weight
                        filtered_right = self.latest_right_weight
                        raw_left = round(left_kg - self.zero_offset, 1)
                        raw_right = round(right_kg - self.zero_offset, 1)
                        raw_combined = round(combined - self.zero_offset, 1)
                        self._excel_rows.append([
                            self.record_timestamp_ms,
                            raw_left, raw_right, raw_combined,
                            filtered_left, filtered_right, filtered_combined])
                        self.record_timestamp_ms += 1
            else:
                for i in range(0, DATA_FIELD_LEN, 3):
                    weight_kg = self._three_bytes_to_weight_kg(
                        frame[4+i], frame[5+i], frame[6+i])
                    if weight_kg is None:
                        continue
                    self._push_point(weight_kg)
                    if self.recording:
                        zeroed     = self.latest_filtered_weight
                        zeroed_raw = round(weight_kg - self.zero_offset, 1)
                        self._excel_rows.append([
                            self.record_timestamp_ms, zeroed_raw, zeroed])
                        self.record_timestamp_ms += 1

    def _clean_three_point(self, raw_buf, res_buf, value):
        raw_buf.append(value)
        if len(raw_buf) < 3:
            cleaned = value
            res_buf.append(cleaned)
            return cleaned, False

        b_mid = raw_buf[-2]
        b_cur = raw_buf[-1]
        f_prev = res_buf[-2] if len(res_buf) >= 2 else res_buf[-1]
        interp = (f_prev + b_cur) / 2.0
        if (abs(b_mid - f_prev) > self.THRESHOLD_DETECT and
                abs(interp - b_mid) > self.THRESHOLD_KEEP):
            cleaned = interp
        else:
            cleaned = b_mid

        res_buf.append(cleaned)
        if len(res_buf) > 2:
            res_buf.pop(0)
        if len(raw_buf) > 3:
            raw_buf.pop(0)
        return cleaned, True

    def _push_point(self, value, left_value=None, right_value=None):
        """三点滑动窗口数据清洗，与 win.py 逻辑完全一致。"""
        combined_cleaned, combined_ready = self._clean_three_point(
            self._raw_buf, self._res_buf, value)

        left_filtered = None
        right_filtered = None
        if left_value is not None:
            left_cleaned, _ = self._clean_three_point(
                self._left_raw_buf, self._left_res_buf, left_value)
            left_filtered = round(left_cleaned - self.zero_offset, 1)
        if right_value is not None:
            right_cleaned, _ = self._clean_three_point(
                self._right_raw_buf, self._right_res_buf, right_value)
            right_filtered = round(right_cleaned - self.zero_offset, 1)

        if not combined_ready:
            return

        combined_value = round(combined_cleaned - self.zero_offset, 1)
        self.data_count += 1
        self.data_sequence += 1
        with self.data_lock:
            self.chart_data_x.append(self.data_sequence)
            self.chart_data_y.append(combined_value)
            self.chart_data_left_y.append(left_filtered if left_filtered is not None else combined_value)
            self.chart_data_right_y.append(right_filtered if right_filtered is not None else 0.0)
            self.latest_filtered_weight = combined_value
            self.latest_left_weight = left_filtered
            self.latest_right_weight = right_filtered

    # ═══════════════════════════════════════════════════════════ 录制
    def toggle_recording(self):
        if self.recording:
            self.stop_recording()
        else:
            self.start_recording()

    def start_recording(self):
        if self.recording:
            return
        self.recording = True
        self.record_timestamp_ms = 1
        self.filename = time.strftime("测力计%Y%m%d_%H%M%S.xlsx")
        with self.data_lock:
            dual = self.is_dual_mode
        self._excel_dual = dual
        self._excel_rows = []
        if dual:
            self._excel_header = ['时间戳(ms)', '主设备原始重量(kg)', '副设备原始重量(kg)', '原始合力(kg)', '主设备清洗重量(kg)', '副设备清洗重量(kg)', '清洗合力(kg)']
        else:
            self._excel_header = ['时间戳(ms)', '重量(kg)', '滤波重量(kg)']
        self.record_button.config(text="停止记录 (r)", bg="#e53935")
        self.status_label.config(text="状态：记录数据中", fg="#b71c1c")
        print("开始记录数据")

    def stop_recording(self):
        if not self.recording:
            return
        self.recording = False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "测力计数据"
            hdr_font  = Font(bold=True, color="FFFFFF")
            hdr_fill  = PatternFill(fill_type="solid", fgColor="1A6FD4")
            hdr_align = Alignment(horizontal="center")
            for col, h in enumerate(self._excel_header, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = hdr_align
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)
                ].width = max(14, len(h) * 2)
            for row in self._excel_rows:
                ws.append(row)
            wb.save(self.filename)
            print(f"已导出 Excel: {self.filename}")
        except Exception as e:
            print(f"导出 Excel 失败: {e}")
            messagebox.showerror("导出失败", str(e))
        self._excel_rows = []
        self.record_button.config(text="开始记录 (r)", bg="#2196f3")
        if self.running:
            self.status_label.config(text="状态：接收数据中", fg="#2e7d32")
        print("停止记录并导出Excel")

    # ═══════════════════════════════════════════════════════════ 关闭
    def _on_close(self):
        self.running = False
        self._stop_ble()
        self.root.after(300, self.root.destroy)


if __name__ == '__main__':
    root = tk.Tk()
    app = BLESensorApp(root)
    root.mainloop()
